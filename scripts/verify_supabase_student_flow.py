#!/usr/bin/env python3
import json
import os
import time
import urllib.error
import urllib.request

from openpyxl import load_workbook


def request_json(method, url, headers, payload=None):
    body = None if payload is None else json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            raw = response.read().decode("utf-8")
            return json.loads(raw) if raw else None
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {error.code}: {detail[:500]}") from None


def load_public_config(path):
    values = {}
    with open(path, encoding="utf-8") as handle:
        for line in handle:
            if "=" in line and not line.lstrip().startswith("#"):
                key, value = line.rstrip().split("=", 1)
                values[key] = value
    return values


def first_student(workbook_path):
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    rows = list(workbook["Alumnos"].iter_rows(values_only=True))
    headers = list(rows[0])
    row = rows[1]
    email = str(row[headers.index("email")]).strip().lower()
    raw_pin = row[headers.index("pin")]
    pin = str(int(raw_pin)) if isinstance(raw_pin, (int, float)) else str(raw_pin).strip()
    return email, pin


def main():
    config = load_public_config(".env.local")
    base_url = config["SUPABASE_URL"].rstrip("/")
    publishable_key = config["SUPABASE_PUBLISHABLE_KEY"]
    email, pin = first_student(".supabase/LenguArcade_DB.xlsx")
    public_headers = {
        "apikey": publishable_key,
        "Authorization": f"Bearer {publishable_key}",
        "Content-Type": "application/json",
    }

    started = time.perf_counter()
    auth = request_json("POST", f"{base_url}/auth/v1/signup", public_headers, {"data": {}})
    auth_ms = round((time.perf_counter() - started) * 1000)
    token = auth.get("access_token")
    if not token:
        raise RuntimeError("No se creo la sesion anonima.")

    user_headers = {
        "apikey": publishable_key,
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    before = request_json(
        "GET",
        f"{base_url}/rest/v1/profiles?select=id&limit=2",
        user_headers,
    )
    if before:
        raise RuntimeError("Una sesion sin PIN pudo leer perfiles.")

    started = time.perf_counter()
    login = request_json(
        "POST",
        f"{base_url}/functions/v1/pin-login",
        user_headers,
        {"email": email, "pin": pin},
    )
    login_ms = round((time.perf_counter() - started) * 1000)
    if not login.get("ok") or login.get("profile", {}).get("email") != email:
        raise RuntimeError("El login por PIN no devolvio el perfil esperado.")

    started = time.perf_counter()
    dashboard = request_json(
        "POST",
        f"{base_url}/functions/v1/student-dashboard",
        user_headers,
        {},
    )
    dashboard_ms = round((time.perf_counter() - started) * 1000)
    if not dashboard.get("ok") or dashboard.get("student", {}).get("email") != email:
        raise RuntimeError("El panel no corresponde al alumno autenticado.")
    if len(dashboard.get("games", [])) < 2:
        raise RuntimeError("El panel no contiene el catalogo de juegos.")

    after = request_json(
        "GET",
        f"{base_url}/rest/v1/profiles?select=id&limit=2",
        user_headers,
    )
    if len(after or []) != 1:
        raise RuntimeError("RLS no limito la lectura al unico perfil autenticado.")

    print(json.dumps({
        "ok": True,
        "profilesVisibleBeforePin": len(before or []),
        "profilesVisibleAfterPin": len(after or []),
        "gamesLoaded": len(dashboard.get("games", [])),
        "eventsLoaded": len(dashboard.get("events", [])),
        "achievementsLoaded": len(dashboard.get("achievements", [])),
        "timingsMs": {
            "anonymousAuth": auth_ms,
            "pinLogin": login_ms,
            "dashboard": dashboard_ms,
        },
    }, indent=2))


if __name__ == "__main__":
    main()
