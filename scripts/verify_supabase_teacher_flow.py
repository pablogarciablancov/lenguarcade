#!/usr/bin/env python3
import json
import os
import sys
import time
import urllib.error
import urllib.request

from openpyxl import load_workbook


def request(url, key, path, payload, access_token=None):
    body = json.dumps(payload).encode("utf-8")
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {access_token or key}",
        "Content-Type": "application/json",
    }
    started = time.perf_counter()
    call = urllib.request.Request(
        f"{url.rstrip('/')}{path}",
        data=body,
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(call, timeout=60) as response:
            data = json.loads(response.read().decode("utf-8") or "{}")
            return data, round((time.perf_counter() - started) * 1000)
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {error.code} en {path}: {detail[:500]}") from None


def teacher_password(workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows = list(workbook["Config"].iter_rows(values_only=True))
    headers = [str(value or "") for value in rows[0]]
    key_index = headers.index("key")
    value_index = headers.index("value")
    for row in rows[1:]:
        if str(row[key_index] or "").strip() == "TEACHER_PASSWORD":
            return str(row[value_index] or "")
    raise RuntimeError("No se encontro TEACHER_PASSWORD en el libro de migracion.")


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Uso: verify_supabase_teacher_flow.py RUTA_XLSX")
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_PUBLISHABLE_KEY", "").strip()
    if not url or not key:
        raise SystemExit("Faltan SUPABASE_URL o SUPABASE_PUBLISHABLE_KEY.")

    password = teacher_password(sys.argv[1])
    auth, auth_ms = request(url, key, "/auth/v1/signup", {"data": {}})
    access_token = auth.get("access_token")
    if not access_token:
        raise RuntimeError("Supabase Auth no devolvio una sesion anonima.")

    login, login_ms = request(
        url,
        key,
        "/functions/v1/teacher-login",
        {"password": password},
        access_token,
    )
    if not login.get("ok"):
        raise RuntimeError("El login del profesor no ha sido aceptado.")

    dashboard, dashboard_ms = request(
        url,
        key,
        "/functions/v1/teacher-dashboard",
        {},
        access_token,
    )
    students = dashboard.get("students") or []
    if not students:
        raise RuntimeError("El panel del profesor no ha devuelto alumnos.")

    detail, detail_ms = request(
        url,
        key,
        "/functions/v1/teacher-student-detail",
        {"studentId": students[0]["studentId"]},
        access_token,
    )
    if not detail.get("student"):
        raise RuntimeError("La ficha detallada no ha devuelto el alumno.")

    grades, grades_ms = request(
        url,
        key,
        "/functions/v1/classroom-sync",
        {"action": "grade-export"},
        access_token,
    )
    print(json.dumps({
        "ok": True,
        "students": len(students),
        "classes": len(dashboard.get("classes") or []),
        "games": len(dashboard.get("games") or []),
        "detailProgress": len(detail.get("progress") or []),
        "gradeCoursesReady": len(grades.get("courses") or []),
        "timingsMs": {
            "anonymousAuth": auth_ms,
            "teacherLogin": login_ms,
            "dashboard": dashboard_ms,
            "studentDetail": detail_ms,
            "gradeExport": grades_ms,
        },
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
