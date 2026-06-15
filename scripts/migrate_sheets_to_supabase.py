#!/usr/bin/env python3
import argparse
import json
import math
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timezone

from openpyxl import load_workbook


def json_value(value, fallback=None):
    if isinstance(value, (dict, list)):
        return value
    if value in (None, ""):
        return {} if fallback is None else fallback
    try:
        return json.loads(str(value))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {} if fallback is None else fallback


def iso_value(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.isoformat()
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=timezone.utc).isoformat()
    return str(value)


def integer(value, default=0):
    try:
        number = float(value)
        return default if math.isnan(number) else int(number)
    except (TypeError, ValueError):
        return default


def decimal(value, default=0):
    try:
        number = float(value)
        return default if math.isnan(number) else number
    except (TypeError, ValueError):
        return default


def pin_value(value):
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return str(value or "").strip()


def boolean(value, default=True):
    if value in (None, ""):
        return default
    return str(value).strip().lower() not in {"false", "0", "no"}


def chunks(values, size=100):
    for index in range(0, len(values), size):
        yield values[index:index + size]


def read_table(workbook, sheet_name):
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for values in rows[1:]:
        if not any(value not in (None, "") for value in values):
            continue
        result.append({
            header: values[index] if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        })
    return result


class SupabaseRest:
    def __init__(self, base_url, secret_key, dry_run=False):
        self.base_url = base_url.rstrip("/")
        self.secret_key = secret_key
        self.dry_run = dry_run

    def request(self, method, path, payload=None, prefer=None):
        if self.dry_run and method not in {"GET", "HEAD"}:
            return []
        body = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
        headers = {
            "apikey": self.secret_key,
            "Authorization": f"Bearer {self.secret_key}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        }
        if prefer:
            headers["Prefer"] = prefer
        request = urllib.request.Request(
            f"{self.base_url}{path}",
            data=body,
            headers=headers,
            method=method,
        )
        try:
            with urllib.request.urlopen(request, timeout=90) as response:
                raw = response.read().decode("utf-8")
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as error:
            detail = error.read().decode("utf-8", errors="replace")
            raise RuntimeError(
                f"Supabase respondio HTTP {error.code} en {method} {path}: {detail[:800]}"
            ) from None

    def select(self, table, query=""):
        suffix = f"?{query}" if query else ""
        return self.request("GET", f"/rest/v1/{table}{suffix}") or []

    def insert(self, table, rows, return_rows=False):
        if not rows:
            return []
        prefer = "return=representation" if return_rows else "return=minimal"
        return self.request("POST", f"/rest/v1/{table}", rows, prefer=prefer) or []

    def upsert(self, table, rows, conflict, return_rows=False):
        if not rows:
            return []
        query = urllib.parse.urlencode({"on_conflict": conflict}, safe=",")
        returning = "representation" if return_rows else "minimal"
        prefer = f"resolution=merge-duplicates,return={returning}"
        return self.request("POST", f"/rest/v1/{table}?{query}", rows, prefer=prefer) or []

    def patch(self, table, filters, payload):
        query = urllib.parse.urlencode(filters, safe=".,()*")
        return self.request("PATCH", f"/rest/v1/{table}?{query}", payload, prefer="return=minimal")

    def rpc(self, name, payload):
        return self.request("POST", f"/rest/v1/rpc/{name}", payload)


def avatar_value(raw):
    if raw in (None, ""):
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        parsed = json.loads(str(raw))
        if isinstance(parsed, dict):
            return parsed
    except (TypeError, ValueError, json.JSONDecodeError):
        pass
    return {"legacy": str(raw)}


def latest_rows(rows, key_builder):
    selected = {}
    for row in rows:
        key = key_builder(row)
        timestamp = str(row.get("updatedAt") or row.get("lastActivity") or "")
        current = selected.get(key)
        if current is None or timestamp >= current[0]:
            selected[key] = (timestamp, row)
    return [entry[1] for entry in selected.values()]


def extract_save(raw_data):
    if not isinstance(raw_data, dict):
        return None
    candidates = [
        raw_data.get("save"),
        (raw_data.get("rawGameData") or {}).get("save")
        if isinstance(raw_data.get("rawGameData"), dict) else None,
    ]
    for candidate in candidates:
        if isinstance(candidate, dict) and candidate:
            return candidate
    return None


def migrate(workbook_path, api, dry_run=False):
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    table_names = [
            "Clases",
            "Alumnos",
            "Juegos",
            "ProgresoJuegos",
            "Eventos",
            "Logros",
            "Misiones",
            "Evaluaciones",
            "Errores",
        ]
    if "Config" in workbook.sheetnames:
        table_names.append("Config")
    tables = {
        name: read_table(workbook, name)
        for name in table_names
    }

    organizations = api.select(
        "organizations",
        urllib.parse.urlencode({"select": "id", "name": "eq.LenguArcade"}),
    )
    if organizations:
        organization_id = organizations[0]["id"]
    else:
        created = api.insert("organizations", [{
            "name": "LenguArcade",
            "email_domain": "alumno.fomento.edu",
            "timezone": "Europe/Madrid",
        }], return_rows=True)
        organization_id = created[0]["id"] if created else "dry-run-organization"

    game_rows = [{
        "id": str(row.get("gameId") or ""),
        "name": str(row.get("nombre") or row.get("gameId") or ""),
        "subtitle": str(row.get("subtitulo") or ""),
        "category": str(row.get("categoria") or ""),
        "status": str(row.get("estado") or "beta"),
        "sort_order": integer(row.get("orden")),
        "color": str(row.get("color") or ""),
        "icon": str(row.get("icono") or ""),
        "url": str(row.get("url") or ""),
        "banner": str(row.get("banner") or ""),
        "active": boolean(row.get("activo")),
    } for row in tables["Juegos"] if row.get("gameId")]
    for batch in chunks(game_rows):
        api.upsert("games", batch, "id")

    existing_classrooms = api.select(
        "classrooms",
        urllib.parse.urlencode({
            "select": "id,legacy_class_code",
            "organization_id": f"eq.{organization_id}",
            "source": "eq.sheets",
        }),
    )
    classroom_by_code = {
        str(row.get("legacy_class_code") or ""): row["id"]
        for row in existing_classrooms
        if row.get("legacy_class_code")
    }
    for row in tables["Clases"]:
        code = str(row.get("classCode") or "").strip()
        if not code:
            continue
        payload = {
            "organization_id": organization_id,
            "legacy_class_code": code,
            "name": str(row.get("nombreVisible") or code),
            "section": str(row.get("linea") or ""),
            "course_state": "ACTIVE" if boolean(row.get("activa")) else "ARCHIVED",
            "active": boolean(row.get("activa")),
            "source": "sheets",
        }
        if code in classroom_by_code:
            api.patch("classrooms", {"id": f"eq.{classroom_by_code[code]}"}, payload)
        else:
            created = api.insert("classrooms", [payload], return_rows=True)
            if created:
                classroom_by_code[code] = created[0]["id"]

    students_by_email = defaultdict(list)
    for row in tables["Alumnos"]:
        email = str(row.get("email") or "").strip().lower()
        if email:
            students_by_email[email].append(row)

    profile_rows = []
    for email, aliases in students_by_email.items():
        primary = aliases[0]
        profile_rows.append({
            "organization_id": organization_id,
            "email": email,
            "first_name": str(primary.get("nombre") or ""),
            "last_name": str(primary.get("apellidos") or ""),
            "role": "student",
            "avatar": avatar_value(primary.get("avatar")),
            "active": any(boolean(row.get("activo")) for row in aliases),
            "source": "sheets",
            "last_login_at": max(
                (iso_value(row.get("ultimaSesion")) or "" for row in aliases),
                default="",
            ) or None,
        })
    for batch in chunks(profile_rows):
        api.upsert("profiles", batch, "organization_id,email")

    profiles = api.select(
        "profiles",
        urllib.parse.urlencode({
            "select": "id,email",
            "organization_id": f"eq.{organization_id}",
        }),
    )
    profile_by_email = {str(row["email"]).lower(): row["id"] for row in profiles}

    alias_rows = []
    alias_to_profile = {}
    enrollment_rows = []
    pin_by_profile = {}
    duplicate_emails = 0
    for email, aliases in students_by_email.items():
        profile_id = profile_by_email.get(email)
        if not profile_id:
            continue
        if len(aliases) > 1:
            duplicate_emails += 1
        pin = pin_value(aliases[0].get("pin"))
        if pin:
            pin_by_profile[profile_id] = pin
        for row in aliases:
            legacy_id = str(row.get("studentId") or "").strip()
            if legacy_id:
                alias_to_profile[legacy_id] = profile_id
                alias_rows.append({
                    "organization_id": organization_id,
                    "source": "sheets",
                    "alias": legacy_id,
                    "profile_id": profile_id,
                })
            classroom_id = classroom_by_code.get(str(row.get("clase") or ""))
            if classroom_id:
                enrollment_rows.append({
                    "classroom_id": classroom_id,
                    "profile_id": profile_id,
                    "active": boolean(row.get("activo")),
                })
    for batch in chunks(alias_rows):
        api.upsert("profile_aliases", batch, "organization_id,source,alias")
    for batch in chunks(latest_rows(
        enrollment_rows,
        lambda row: (row["classroom_id"], row["profile_id"]),
    )):
        api.upsert("classroom_enrollments", batch, "classroom_id,profile_id")
    for profile_id, pin in pin_by_profile.items():
        if not dry_run:
            api.rpc("set_profile_pin", {
                "target_profile_id": profile_id,
                "plain_pin": pin,
            })

    teacher_password = ""
    for row in tables.get("Config", []):
        if str(row.get("key") or "").strip() == "TEACHER_PASSWORD":
            teacher_password = str(row.get("value") or "").strip()
            break
    teacher_profile_id = None
    if teacher_password:
        teacher_rows = api.upsert("profiles", [{
            "organization_id": organization_id,
            "email": "profesor@lenguarcade.local",
            "first_name": "Profesor",
            "last_name": "LenguArcade",
            "role": "teacher",
            "active": True,
            "source": "sheets",
        }], "organization_id,email", return_rows=True)
        if teacher_rows:
            teacher_profile_id = teacher_rows[0]["id"]
            teacher_links = [{
                "classroom_id": classroom_id,
                "profile_id": teacher_profile_id,
                "is_owner": True,
            } for classroom_id in classroom_by_code.values()]
            for batch in chunks(teacher_links):
                api.upsert("classroom_teachers", batch, "classroom_id,profile_id")
            if not dry_run:
                api.rpc("set_teacher_password", {
                    "target_profile_id": teacher_profile_id,
                    "plain_password": teacher_password,
                })

    progress_source = latest_rows(
        [
            row for row in tables["ProgresoJuegos"]
            if str(row.get("studentId") or "") in alias_to_profile and row.get("gameId")
        ],
        lambda row: (
            alias_to_profile[str(row.get("studentId"))],
            str(row.get("gameId")),
        ),
    )
    progress_rows = []
    save_rows = []
    for row in progress_source:
        profile_id = alias_to_profile[str(row.get("studentId"))]
        game_id = str(row.get("gameId"))
        raw_data = json_value(row.get("rawJson"))
        progress_rows.append({
            "profile_id": profile_id,
            "game_id": game_id,
            "xp": max(0, integer(row.get("xp"))),
            "level": max(1, integer(row.get("nivel"), 1)),
            "percentage": max(0, min(100, decimal(row.get("percentage")))),
            "accuracy": max(0, min(100, decimal(row.get("accuracy")))),
            "attempts": max(0, integer(row.get("attempts"))),
            "successes": max(0, integer(row.get("successes"))),
            "errors": max(0, integer(row.get("errors"))),
            "streak": max(0, integer(row.get("streak"))),
            "sessions": max(0, integer(row.get("sessions"))),
            "achievements_count": max(0, integer(row.get("achievementsCount"))),
            "missions_completed": max(0, integer(row.get("missionsCompleted"))),
            "feathers": max(0, integer(row.get("plumas"))),
            "raw_data": raw_data,
            "last_activity_at": iso_value(row.get("lastActivity")),
        })
        save_data = extract_save(raw_data)
        if save_data:
            save_rows.append({
                "profile_id": profile_id,
                "game_id": game_id,
                "slot": "main",
                "revision": 1,
                "save_data": save_data,
                "saved_at": iso_value(row.get("updatedAt")) or iso_value(row.get("lastActivity")),
            })
    for batch in chunks(progress_rows):
        api.upsert("game_progress", batch, "profile_id,game_id")
    for batch in chunks(save_rows):
        api.upsert("game_saves", batch, "profile_id,game_id,slot")

    event_rows = []
    for index, row in enumerate(tables["Eventos"]):
        legacy_id = str(row.get("studentId") or "")
        profile_id = alias_to_profile.get(legacy_id)
        game_id = str(row.get("gameId") or "")
        if not profile_id or not game_id:
            continue
        result_id = str(row.get("eventId") or f"legacy-event-{index}")
        accuracy = row.get("accuracy")
        event_rows.append({
            "result_id": result_id,
            "profile_id": profile_id,
            "game_id": game_id,
            "event_type": str(row.get("eventType") or "legacy_event"),
            "xp_delta": integer(row.get("xpDelta")),
            "feathers_delta": integer(row.get("plumasDelta")),
            "accuracy": None if accuracy in (None, "") else max(0, min(100, decimal(accuracy))),
            "details": json_value(row.get("detailsJson")),
            "occurred_at": iso_value(row.get("timestamp")) or datetime.now(timezone.utc).isoformat(),
        })
    for batch in chunks(event_rows):
        api.upsert("game_events", batch, "profile_id,game_id,result_id")

    definitions = {}
    player_achievements = []
    for row in tables["Logros"]:
        legacy_id = str(row.get("studentId") or "")
        profile_id = alias_to_profile.get(legacy_id)
        game_id = str(row.get("gameId") or "")
        achievement_id = str(row.get("achievementId") or "")
        if not profile_id or not game_id or not achievement_id:
            continue
        definitions[(game_id, achievement_id)] = {
            "game_id": game_id,
            "id": achievement_id,
            "title": str(row.get("title") or achievement_id),
            "description": str(row.get("description") or ""),
            "xp_reward": integer(row.get("xpReward")),
            "hidden": False,
        }
        player_achievements.append({
            "profile_id": profile_id,
            "game_id": game_id,
            "achievement_id": achievement_id,
            "unlocked_at": iso_value(row.get("unlockedAt")) or datetime.now(timezone.utc).isoformat(),
        })
    for batch in chunks(list(definitions.values())):
        api.upsert("achievement_definitions", batch, "game_id,id")
    for batch in chunks(latest_rows(
        player_achievements,
        lambda row: (row["profile_id"], row["game_id"], row["achievement_id"]),
    )):
        api.upsert("player_achievements", batch, "profile_id,game_id,achievement_id")

    mission_rows = [{
        "id": str(row.get("missionId") or ""),
        "title": str(row.get("title") or ""),
        "description": str(row.get("description") or ""),
        "game_id": str(row.get("gameId") or "general"),
        "mission_type": str(row.get("type") or ""),
        "target": decimal(row.get("target")),
        "reward_xp": integer(row.get("rewardXp")),
        "reward_feathers": integer(row.get("rewardPlumas")),
        "active_from": iso_value(row.get("activeFrom")),
        "active_to": iso_value(row.get("activeTo")),
        "active": boolean(row.get("isActive")),
    } for row in tables["Misiones"] if row.get("missionId")]
    for batch in chunks(mission_rows):
        api.upsert("mission_definitions", batch, "id")

    evaluation_rows = []
    for row in tables["Evaluaciones"]:
        profile_id = alias_to_profile.get(str(row.get("studentId") or ""))
        if not profile_id:
            continue
        evaluation_rows.append({
            "profile_id": profile_id,
            "classroom_id": classroom_by_code.get(str(row.get("classCode") or "")),
            "scope": str(row.get("scope") or "general"),
            "game_id": str(row.get("gameId") or "general"),
            "score": decimal(row.get("score")),
            "breakdown": json_value(row.get("breakdownJson")),
            "updated_at": iso_value(row.get("updatedAt")) or datetime.now(timezone.utc).isoformat(),
        })
    for batch in chunks(evaluation_rows):
        api.upsert("evaluations", batch, "profile_id,scope,game_id")

    error_rows = []
    for row in tables["Errores"]:
        profile_id = alias_to_profile.get(str(row.get("studentId") or ""))
        game_id = str(row.get("gameId") or "")
        if not profile_id or not game_id:
            continue
        error_rows.append({
            "profile_id": profile_id,
            "game_id": game_id,
            "skill": str(row.get("skill") or ""),
            "error_type": str(row.get("errorType") or ""),
            "error_count": max(1, integer(row.get("count"), 1)),
            "details": json_value(row.get("detailsJson")),
            "occurred_at": iso_value(row.get("timestamp")) or datetime.now(timezone.utc).isoformat(),
        })
    for batch in chunks(error_rows):
        api.insert("game_errors", batch)

    summary = {
        "dryRun": dry_run,
        "classes": len(classroom_by_code),
        "sourceStudentRows": len(tables["Alumnos"]),
        "profiles": len(profile_rows),
        "duplicateEmailsMerged": duplicate_emails,
        "aliases": len(alias_rows),
        "progressRows": len(progress_rows),
        "saves": len(save_rows),
        "events": len(event_rows),
        "achievementDefinitions": len(definitions),
        "playerAchievements": len(player_achievements),
        "missions": len(mission_rows),
        "evaluations": len(evaluation_rows),
        "errors": len(error_rows),
        "teacherConfigured": bool(teacher_profile_id),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    url = os.environ.get("SUPABASE_URL", "").strip()
    secret = os.environ.get("SUPABASE_SECRET_KEY", "").strip()
    if not url or not secret:
        raise SystemExit("Faltan SUPABASE_URL o SUPABASE_SECRET_KEY.")
    migrate(args.workbook, SupabaseRest(url, secret, args.dry_run), args.dry_run)


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"Error de migracion: {error}", file=sys.stderr)
        raise
